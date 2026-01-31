import pandas as pd
from datetime import datetime, time
import os
import shutil
import glob
import unicodedata
import re
from colorama import init, Fore, Style

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Inicializar colorama
init(autoreset=True)

def remover_acentos(texto):
    """Remove acentos de um texto"""
    nfd = unicodedata.normalize('NFD', texto)
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')

def _normalizar_com_mapa(texto):
    """
    Normaliza removendo acentos e retorna:
    - texto_normalizado
    - mapa_indices: lista onde cada índice do texto normalizado aponta para o índice correspondente no texto original
    """
    normalizado_chars = []
    mapa_indices = []
    for idx, ch in enumerate(texto):
        ch_normalizado = remover_acentos(ch)
        if not ch_normalizado:
            continue
        for _ in ch_normalizado:
            normalizado_chars.append(_.upper())
            mapa_indices.append(idx)
    return ''.join(normalizado_chars), mapa_indices

def _extrair_secao_texto(conteudo, cabecalhos_secao, cabecalhos_proxima_secao):
    """
    Extrai uma seção do texto usando busca sem acentos, preservando os índices do texto original.
    Retorna a substring original (sem o cabeçalho).
    """
    if not conteudo:
        return ""

    conteudo_normalizado, mapa_indices = _normalizar_com_mapa(conteudo)
    if not conteudo_normalizado:
        return ""

    cabecalhos_norm = [remover_acentos(cab.upper()) for cab in cabecalhos_secao]
    proximos_norm = [remover_acentos(cab.upper()) for cab in cabecalhos_proxima_secao]

    inicio_norm = -1
    cabecalho_encontrado = ""
    for cab in cabecalhos_norm:
        match = re.search(re.escape(cab), conteudo_normalizado)
        if match:
            inicio_norm = match.start() + len(cab)
            cabecalho_encontrado = cab
            break

    if inicio_norm == -1:
        return ""

    fim_norm = -1
    if proximos_norm:
        match_fim = re.search('|'.join(re.escape(p) for p in proximos_norm), conteudo_normalizado[inicio_norm:])
        if match_fim:
            fim_norm = inicio_norm + match_fim.start()

    if inicio_norm >= len(mapa_indices):
        return ""

    inicio_orig = mapa_indices[inicio_norm]
    if fim_norm != -1 and fim_norm - 1 < len(mapa_indices):
        fim_orig = mapa_indices[fim_norm - 1] + 1
        return conteudo[inicio_orig:fim_orig]
    return conteudo[inicio_orig:]

def _extrair_secao_por_linha(conteudo, cabecalhos_secao, cabecalhos_proxima_secao):
    """
    Extrai uma seção procurando cabeçalhos em linhas isoladas (sem texto após ':').
    Usa comparação sem acentos. Retorna a substring original (sem o cabeçalho).
    """
    if not conteudo:
        return ""

    cabecalhos_norm = {remover_acentos(cab.upper()).strip() for cab in cabecalhos_secao}
    proximos_norm = {remover_acentos(cab.upper()).strip() for cab in cabecalhos_proxima_secao}

    inicio_idx = None
    fim_idx = None
    offset = 0

    for linha in conteudo.splitlines(keepends=True):
        linha_norm = remover_acentos(linha.upper()).strip()
        if inicio_idx is None and linha_norm in cabecalhos_norm:
            inicio_idx = offset + len(linha)
        elif inicio_idx is not None and linha_norm in proximos_norm:
            fim_idx = offset
            break
        offset += len(linha)

    if inicio_idx is None:
        return ""
    if fim_idx is None:
        return conteudo[inicio_idx:]
    return conteudo[inicio_idx:fim_idx]


def extrair_motoristas_atraso(arquivo_excel, coluna_motorista, coluna_apresenta, coluna_escala):
    """
    Extrai motoristas em atraso que possuem ANOTAÇÕES (comentários do Excel) na coluna APRESENTA
    Retorna string formatada: MOTORISTA - ESCALA: HH:MM - ANOTAÇÃO
    """
    motoristas_atraso = ""
    
    if not OPENPYXL_AVAILABLE:
        return motoristas_atraso
    
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        
        # Encontrar índices das colunas
        col_motorista_idx = None
        col_apresenta_idx = None
        col_escala_idx = None
        
        for cell in ws[1]:
            if cell.value:
                if 'MOTORISTA' in str(cell.value).upper():
                    col_motorista_idx = cell.column
                if 'APRESENTA' in str(cell.value).upper():
                    col_apresenta_idx = cell.column
                if 'ESCALA' in str(cell.value).upper():
                    col_escala_idx = cell.column
        
        if not col_apresenta_idx or not col_motorista_idx or not col_escala_idx:
            wb.close()
            return motoristas_atraso
        
        # Iterar pelas linhas procurando por comentários em APRESENTA
        for row_num in range(2, ws.max_row + 1):
            cell_apresenta = ws.cell(row=row_num, column=col_apresenta_idx)
            
            # Verificar se a célula tem comentário/anotação
            if cell_apresenta.comment:
                anotacao_texto = cell_apresenta.comment.text
                
                # Obter dados da mesma linha
                cell_motorista = ws.cell(row=row_num, column=col_motorista_idx)
                cell_escala = ws.cell(row=row_num, column=col_escala_idx)
                
                motorista_val = cell_motorista.value
                escala_val = cell_escala.value
                
                if motorista_val and anotacao_texto:
                    motorista_str = str(motorista_val).strip()
                    anotacao_str = str(anotacao_texto).strip()
                    
                    # Extrair apenas o corpo da anotação (após os :)
                    if ':' in anotacao_str:
                        anotacao_str = anotacao_str.split(':', 1)[1].strip()
                    
                    # Formatar escala
                    if isinstance(escala_val, datetime):
                        escala_str = escala_val.strftime('%H:%M')
                    elif isinstance(escala_val, time):
                        escala_str = escala_val.strftime('%H:%M')
                    else:
                        escala_str = str(escala_val).strip() if escala_val else ""
                    
                    motoristas_atraso += f"{motorista_str} - ESCALA: {escala_str} - {anotacao_str}\n"
        
        wb.close()
        
    except Exception as e:
        print(f"{Fore.YELLOW}⚠ Erro ao extrair motoristas em atraso: {e}{Style.RESET_ALL}")
    
    return motoristas_atraso


def extrair_placa_de_linha_pavao(linha):
    """
    Extrai placa de uma linha do PAVÃO.
    Aceita:
    - "PLACA: XXXXXX"
    - linha iniciando com XXXXXX (6 caracteres)
    Retorna a placa limpa (sem hífen) ou "".
    """
    if not linha:
        return ""

    # Prioriza padrão "PLACA: "
    match = re.search(r'PLACA:\s*([A-Z0-9\-]{6,7})', linha, re.IGNORECASE)
    if match:
        return match.group(1).strip().upper().replace('-', '')

    # Fallback: placa em qualquer lugar da linha (6 caracteres)
    match_qualquer = re.search(r'(?<![A-Z0-9\-])([A-Z0-9\-]{6,7})(?![A-Z0-9\-])', linha, re.IGNORECASE)
    if match_qualquer:
        return match_qualquer.group(1).strip().upper().replace('-', '')

    return ""

def extrair_placas_de_pavao(texto_pavao):
    """
    Extrai placas do texto de PAVÃO.
    Procura por "PLACA: XXXXXX" ou linha iniciando com XXXXXX.
    Retorna lista de placas limpas (sem hífen).
    """
    placas = []
    if not texto_pavao:
        return placas

    for linha in texto_pavao.splitlines():
        placa = extrair_placa_de_linha_pavao(linha)
        if placa:
            placas.append(placa)

    return placas

def processar_pavao_com_destino(pavao_content, df, colunas_comparacao, pavao_count_feito):
    """
    Remove linhas de PAVÃO que existem em DESTINO
    Extrai exatamente 6 caracteres após "PLACA: "
    Compara com o pavao_count_feito (número de OK contados)
    Retorna: (conteúdo atualizado, lista de placas removidas, aviso se houver discrepâncias)
    """
    if not pavao_content:
        return pavao_content, [], ""

    if isinstance(colunas_comparacao, str):
        colunas_comparacao = [colunas_comparacao]

    colunas_validas = [col for col in (colunas_comparacao or []) if col in df.columns]
    if not colunas_validas:
        return pavao_content, [], ""
    
    # Extrair placas do PAVÃO (somente linhas com "PLACA:")
    placas_pavao = extrair_placas_de_pavao(pavao_content)
    total_pavao_no_report = len(placas_pavao)
    if total_pavao_no_report == 0:
        return pavao_content, [], ""
    
    # Extrair placas das colunas de comparação (DESTINO, CAVALO, etc.)
    placas_destino = set()
    pattern_dest = r'([A-Z0-9\-]{6,7})'
    for col in colunas_validas:
        for dest in df[col].dropna():
            dest_str = str(dest).strip().upper()
            matches_dest = re.findall(pattern_dest, dest_str)
            for match in matches_dest:
                placa_limpa = match.replace('-', '')
                placas_destino.add(placa_limpa)
    
    # Remover linhas do PAVÃO cuja placa está em DESTINO (com ou sem "PLACA:")
    placas_removidas = []
    linhas_pavao = pavao_content.strip().split('\n')
    linhas_atualizadas = []
    
    for linha in linhas_pavao:
        placa = extrair_placa_de_linha_pavao(linha)
        if placa and placa in placas_destino:
            placas_removidas.append(placa)
            continue
        linhas_atualizadas.append(linha)
    
    pavao_atualizado = '\n'.join(linhas_atualizadas).strip()
    
    # Avisar sobre discrepâncias
    # Comparar com o pavao_count_feito (OK contados), não com o total no report anterior
    aviso = ""
    if pavao_count_feito != len(placas_removidas):
        aviso = f"\n⚠️  [AVISO] PAVÃO: {str(pavao_count_feito).zfill(2)} PAVÕES foram feitos (OK), mas apenas {str(len(placas_removidas)).zfill(2)} foram encontrados em DESTINO. {str(pavao_count_feito - len(placas_removidas)).zfill(2)} ainda precisam ser removidas manualmente."
    
    return pavao_atualizado, placas_removidas, aviso



def limpar_tela():
    os.system('cls' if os.name == 'nt' else 'clear')

def encontrar_arquivo_escala():
    """Encontra o primeiro arquivo Excel que comece com 'ESCALA' na pasta"""
    arquivos = glob.glob('1.ESCALA-FIM-TURNO/ESCALA*.xlsx')
    if arquivos:
        return arquivos[0]
    return None

def obter_linhas_com_valores_reais(arquivo_excel, nome_coluna_frota):
    """
    Retorna índices das linhas que têm valores reais (não fórmulas) na coluna FROTA
    """
    linhas_reais = set()
    debug_info = []
    
    if not OPENPYXL_AVAILABLE:
        # Se openpyxl não está disponível, retorna todas as linhas (fallback)
        return None
    
    try:
        wb = load_workbook(arquivo_excel)
        ws = wb.active
        
        # Encontrar o índice da coluna FROTA
        coluna_frota_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().upper() == nome_coluna_frota.upper():
                coluna_frota_idx = cell.column
                break
        
        if coluna_frota_idx is None:
            print(f"{Fore.YELLOW}⚠ Coluna {nome_coluna_frota} não encontrada no header{Style.RESET_ALL}")
            return None
        
        print(f"{Fore.CYAN}🔍 Detectando valores reais em FROTA (coluna índice {coluna_frota_idx})...{Style.RESET_ALL}")
        
        # Iterar pelas linhas e verificar se a célula tem fórmula
        for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            cell = row[coluna_frota_idx - 1]
            # Se a célula não começa com "=" (não é fórmula) e tem valor
            if cell.value and not str(cell.value).startswith('='):
                linhas_reais.add(row_num)
                debug_info.append(f"Linha {row_num}: FROTA={cell.value}")
        
        # Mostrar debug info das primeiras 5 linhas encontradas
        if debug_info:
            print(f"{Fore.CYAN}📊 Valores reais encontrados:{Style.RESET_ALL}")
            for info in debug_info[:10]:
                print(f"  {info}")
            if len(debug_info) > 10:
                print(f"  ... e mais {len(debug_info) - 10}")
        
        wb.close()
        return linhas_reais if linhas_reais else None
        
    except Exception as e:
        print(f"{Fore.YELLOW}⚠ Erro ao detectar fórmulas: {e}{Style.RESET_ALL}")
        import traceback
        traceback.print_exc()
        return None

# Function to create the report
def create_report(plano_do_dia, responsavel, aguardando_mdf, aguardando_faturamento):
    # Read the Excel file
    try:
        print(f"{Fore.YELLOW}⏳ Lendo arquivo COLE_AQUI.txt...{Style.RESET_ALL}")
        pavao_content = ""
        pendencias_content = ""
        
        # Tentar ler o arquivo COLE_AQUI.txt
        try:
            with open('2.ULTIMO-REPORT/COLE_AQUI.txt', 'r', encoding='utf-8') as file:
                conteudo = file.read()
                
                # Extrair conteúdo de PAVÃO: (com ou sem acento) - somente cabeçalho em linha isolada
                pavao_raw = _extrair_secao_por_linha(
                    conteudo,
                    ['PAVÃO:', 'PAVAO:'],
                    ['PENDÊNCIAS:', 'PENDENCIAS:']
                )
                if pavao_raw:
                    pavao_content = pavao_raw.strip().upper()
                
                # Extrair conteúdo de PENDÊNCIAS: (com ou sem acento)
                pendencias_raw = _extrair_secao_por_linha(
                    conteudo,
                    ['PENDÊNCIAS:', 'PENDENCIAS:'],
                    ['TROCA DE CAVALO:']
                )
                if pendencias_raw:
                    pendencias_content = pendencias_raw.strip().upper()
                
                print(f"{Fore.CYAN}✓ Arquivo COLE_AQUI.txt lido com sucesso{Style.RESET_ALL}")
        except FileNotFoundError:
            print(f"{Fore.YELLOW}⚠ Arquivo COLE_AQUI.txt não encontrado, usando campos vazios{Style.RESET_ALL}")
        
        print(f"{Fore.YELLOW}⏳ Procurando planilha de escalas...{Style.RESET_ALL}")
        arquivo_escala = encontrar_arquivo_escala()
        
        if not arquivo_escala:
            print(f"{Fore.RED}✗ Nenhum arquivo ESCALA*.xlsx encontrado na pasta 1.ESCALA-FIM-TURNO{Style.RESET_ALL}")
            return
        
        print(f"{Fore.CYAN}📊 Encontrado: {os.path.basename(arquivo_escala)}{Style.RESET_ALL}")
        print(f"{Fore.YELLOW}⏳ Lendo planilha...{Style.RESET_ALL}")
        df = pd.read_excel(arquivo_escala)
        
        # Debug: mostrar nomes das colunas
        print(f"{Fore.CYAN}📋 Colunas encontradas: {list(df.columns)}{Style.RESET_ALL}")
        
        # Process the data as needed
        # For example, let's assume we want to count the number of drivers
        num_motoristas = len(df)
        
        # Analisar colunas VIAGEM e ESCALA para contar enviados
        enviados = 0
        pavao_count = 0
        checkout_count = 0
        saida_itu_dhl = 0
        troca_cavalo_content = ""
        motoristas_atraso_content = ""
        
        # Encontrar todas as colunas que contém "VIAGEM"
        colunas_viagem = [col for col in df.columns if 'VIAGEM' in str(col).upper()]
        
        # Encontrar coluna FROTA e MOTORISTA para TROCA DE CAVALO
        coluna_frota = None
        coluna_motorista = None
        for col in df.columns:
            if 'FROTA' in str(col).upper():
                coluna_frota = col
            if 'MOTORISTA' in str(col).upper():
                coluna_motorista = col
        
        print(f"{Fore.CYAN}📋 Total de colunas encontradas: {len(df.columns)}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}📋 Colunas de viagem: {colunas_viagem}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}📋 Coluna FROTA: {coluna_frota}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}📋 Coluna MOTORISTA: {coluna_motorista}{Style.RESET_ALL}")
        
        # Obter linhas com valores reais (não fórmulas) na coluna FROTA
        linhas_frota_reais = None
        if coluna_frota:
            print(f"{Fore.CYAN}[DEBUG] Chamando obter_linhas_com_valores_reais para coluna: {coluna_frota}{Style.RESET_ALL}")
            linhas_frota_reais = obter_linhas_com_valores_reais(arquivo_escala, coluna_frota)
            print(f"{Fore.CYAN}[DEBUG] Resultado de obter_linhas_com_valores_reais: {linhas_frota_reais}{Style.RESET_ALL}")
            if linhas_frota_reais:
                print(f"{Fore.CYAN}📦 Encontradas {len(linhas_frota_reais)} linhas com valores reais em FROTA{Style.RESET_ALL}")
        
        # Coletar dados de TROCA DE CAVALO ANTES do loop de viagens
        for index, row in df.iterrows():
            # Coletar dados de TROCA DE CAVALO (valores que não são fórmulas)
            if coluna_frota and coluna_motorista and linhas_frota_reais is not None:
                # Usar o índice da linha (index é baseado em 0, mas as linhas reais estão baseadas em 1 do Excel)
                linha_excel = index + 2  # +2 porque Excel começa em 1 e pula o header
                
                if linha_excel in linhas_frota_reais:
                    frota_val = row[coluna_frota]
                    motorista_val = row[coluna_motorista]
                    
                    frota_str = str(frota_val).strip() if pd.notna(frota_val) else ""
                    motorista_str = str(motorista_val).strip() if pd.notna(motorista_val) else ""
                    
                    frota_valida = frota_str and frota_str != 'nan' and frota_str != '-'
                    motorista_valido = motorista_str and motorista_str != 'nan' and motorista_str != '-'
                    if motorista_valido and frota_valida:
                        troca_cavalo_content += f"{motorista_str} - {frota_str}\n"
                        print(f"{Fore.GREEN}  ✓ TROCA DE CAVALO: {motorista_str} - {frota_str}{Style.RESET_ALL}")
            
            # Coletar MOTORISTAS EM ATRASO será feito após a leitura da planilha
        
        if colunas_viagem and 'ESCALA' in df.columns:
            print(f"{Fore.YELLOW}⏳ Analisando viagens...{Style.RESET_ALL}")
            
            # Debug: mostrar primeiras 10 linhas da primeira coluna VIAGEM e ESCALA
            primeira_col_viagem = colunas_viagem[0]
            print(f"\n{Fore.CYAN}📊 Primeiras 10 linhas de {primeira_col_viagem} e ESCALA:{Style.RESET_ALL}")
            for i in range(min(10, len(df))):
                val_viagem = df.iloc[i][primeira_col_viagem]
                val_escala = df.iloc[i]['ESCALA']
                print(f"  Linha {i}: VIAGEM={repr(val_viagem)} | ESCALA={repr(val_escala)}")
            print()
            
            for index, row in df.iterrows():
                # Contar "OK" na coluna VIAGEM (maiúsculo ou minúsculo)
                for col_viagem in colunas_viagem:
                    viagem = str(row[col_viagem]).strip().upper()
                    if viagem == 'OK':
                        pavao_count += 1
                
                # Verificar se alguma das colunas VIAGEM tem "V"
                tem_v = False
                for col_viagem in colunas_viagem:
                    viagem = str(row[col_viagem]).strip().upper()
                    if viagem == 'V':
                        tem_v = True
                        break
                
                if tem_v:
                    escala = row['ESCALA']
                    # Tentar extrair hora do campo ESCALA
                    try:
                        if isinstance(escala, datetime):
                            hora = escala.time()
                        elif isinstance(escala, time):
                            hora = escala
                        elif isinstance(escala, str):
                            # Tentar parsear string no formato "00:00" ou "00:00:00"
                            escala_limpa = escala.strip()
                            hora_partes = escala_limpa.split(':')
                            
                            if len(hora_partes) >= 2:
                                horas = int(hora_partes[0])
                                minutos = int(hora_partes[1])
                                hora = time(horas, minutos)
                            else:
                                print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: V encontrado mas formato de hora inválido: {repr(escala)}{Style.RESET_ALL}")
                                continue
                        elif pd.isna(escala):
                            # Campo vazio/NaN - pula
                            print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: V encontrado mas ESCALA vazia{Style.RESET_ALL}")
                            continue
                        else:
                            print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: V encontrado mas tipo de ESCALA desconhecido: {type(escala)} = {repr(escala)}{Style.RESET_ALL}")
                            continue
                        
                        # Verificar se está entre 00:00 e 05:20
                        if hora >= datetime.strptime('00:00', '%H:%M').time() and hora <= datetime.strptime('05:20', '%H:%M').time():
                            enviados += 1
                            print(f"{Fore.GREEN}  ✓ Linha {index+2}: V com hora {hora} - CONTADO{Style.RESET_ALL}")
                        else:
                            checkout_count += 1
                            print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: V com hora {hora} - FORA DO HORÁRIO (00:00-05:20) - CHECKOUT{Style.RESET_ALL}")
                    except Exception as ex:
                        # Se não conseguir processar a hora, ignora essa linha
                        print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: Erro ao processar - {ex} | ESCALA={repr(escala)}{Style.RESET_ALL}")
                        continue
                
                # Verificar se alguma das colunas VIAGEM tem "SC"
                tem_sc = False
                for col_viagem in colunas_viagem:
                    viagem = str(row[col_viagem]).strip().upper()
                    if viagem == 'SC':
                        tem_sc = True
                        break
                
                if tem_sc:
                    escala = row['ESCALA']
                    # Tentar extrair hora do campo ESCALA
                    try:
                        if isinstance(escala, datetime):
                            hora = escala.time()
                        elif isinstance(escala, time):
                            hora = escala
                        elif isinstance(escala, str):
                            # Tentar parsear string no formato "00:00" ou "00:00:00"
                            escala_limpa = escala.strip()
                            hora_partes = escala_limpa.split(':')
                            
                            if len(hora_partes) >= 2:
                                horas = int(hora_partes[0])
                                minutos = int(hora_partes[1])
                                hora = time(horas, minutos)
                            else:
                                print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: SC encontrado mas formato de hora inválido: {repr(escala)}{Style.RESET_ALL}")
                                continue
                        elif pd.isna(escala):
                            # Campo vazio/NaN - pula
                            print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: SC encontrado mas ESCALA vazia{Style.RESET_ALL}")
                            continue
                        else:
                            print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: SC encontrado mas tipo de ESCALA desconhecido: {type(escala)} = {repr(escala)}{Style.RESET_ALL}")
                            continue
                        
                        # Verificar se está entre 00:00 e 05:20
                        if hora >= datetime.strptime('00:00', '%H:%M').time() and hora <= datetime.strptime('05:20', '%H:%M').time():
                            saida_itu_dhl += 1
                            print(f"{Fore.GREEN}  ✓ Linha {index+2}: SC com hora {hora} - SAÍDA ITU X DHL{Style.RESET_ALL}")
                    except Exception as ex:
                        # Se não conseguir processar a hora, ignora essa linha
                        print(f"{Fore.YELLOW}  ⚠ Linha {index+2}: Erro ao processar SC - {ex} | ESCALA={repr(escala)}{Style.RESET_ALL}")
                        continue
            
        else:
            print(f"{Fore.RED}✗ Colunas necessárias não encontradas!{Style.RESET_ALL}")
            if not colunas_viagem:
                print(f"{Fore.RED}  Nenhuma coluna VIAGEM encontrada{Style.RESET_ALL}")
            if 'ESCALA' not in df.columns:
                print(f"{Fore.RED}  Coluna ESCALA não existe{Style.RESET_ALL}")
        
        print(f"{Fore.GREEN}✓ Planilha carregada com sucesso!{Style.RESET_ALL}")
        print(f"{Fore.CYAN}📦 Enviados encontrados: {str(enviados).zfill(2)}{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}✗ Erro ao ler planilha: {e}{Style.RESET_ALL}")
        return

    # Prepare the report content
    data_atual = datetime.now().strftime('%d/%m')
    
    # Extrair motoristas em atraso (com anotações do Excel na coluna APRESENTA)
    motoristas_atraso_content = extrair_motoristas_atraso(arquivo_escala, coluna_motorista, 'APRESENTA', 'ESCALA')
    if motoristas_atraso_content:
        print(f"{Fore.CYAN}ℹ Motoristas em atraso encontrados com anotações:{Style.RESET_ALL}")
        for linha in motoristas_atraso_content.strip().split('\n'):
            print(f"  - {linha}")
    
    # Processar PAVÃO: remover linhas que correspondem a placas em DESTINO
    colunas_comparacao = [col for col in ['CAVALO', 'DESTINO'] if col in df.columns]
    aviso_pavao = ""
    if colunas_comparacao:
        pavao_content_processado, placas_removidas, aviso_pavao = processar_pavao_com_destino(pavao_content, df, colunas_comparacao, pavao_count)
        if placas_removidas:
            print(f"{Fore.CYAN}ℹ Removidas {str(len(placas_removidas)).zfill(2)} placa(s) do PAVÃO que foram encontradas na escala (CAVALO/DESTINO){Style.RESET_ALL}")
            for placa in placas_removidas:
                print(f"  - {placa}")
        if aviso_pavao:
            print(f"{Fore.YELLOW}{aviso_pavao}{Style.RESET_ALL}")
    else:
        pavao_content_processado = pavao_content
    
    # Adicionar linha PAVAO apenas se existir pelo menos um registro
    pavao_line = f"PAVAO: {str(pavao_count).zfill(2)}\n" if pavao_count > 0 else ""

    saida_itu_dhl_output = str(saida_itu_dhl).zfill(2) if saida_itu_dhl > 0 else ""
    
    report_content = f"""REPORT OPERAÇÃO P2 {data_atual} - {responsavel}

PLANO DO DIA: {plano_do_dia}

ENVIADAS: {str(enviados).zfill(2)}
{pavao_line}

AGUARDANDO MDF: {aguardando_mdf}
AGUARDANDO FATURAMENTO: {aguardando_faturamento}
AGUARDANDO CHECKOUT: {str(checkout_count).zfill(2)}

PAVÃO:

{pavao_content_processado}

PENDÊNCIAS:

{pendencias_content}

TROCA DE CAVALO:

{troca_cavalo_content}

MOVIMENTAÇÕES SÓ CAVALO:

ENTRADA DHL X ITU: 
SAÍDA ITU X DHL: {saida_itu_dhl_output}
SAÍDA ITU x SOROCABA: 
ENTRADA SOROCABA x ITU: 

MOTORISTA EM ATRASO:

{motoristas_atraso_content}

"""

    # Write to a new report file
    timestamp = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    data_sem_ano = datetime.now().strftime('%d-%m')
    arquivo_raiz = 'ULTIMO_RELATORIO.txt'
    arquivo_historico = f'3.HISTORICO-REPORT/REPORT {responsavel} {timestamp}.txt'
    
    # Salvar na raiz
    with open(arquivo_raiz, 'w', encoding='utf-8') as file:
        file.write(report_content)
    
    # Salvar cópia no histórico
    with open(arquivo_historico, 'w', encoding='utf-8') as file:
        file.write(report_content)
    
    # Copiar escala para histórico (sobrescreve se já existe no dia)
    nome_arquivo_escala = os.path.basename(arquivo_escala)
    nome_sem_extensao = os.path.splitext(nome_arquivo_escala)[0]
    arquivo_escala_destino = f'4.HISTORICO-ESCALA/{nome_sem_extensao} {data_sem_ano}.xlsx'
    try:
        shutil.copy2(arquivo_escala, arquivo_escala_destino)
        print(f"{Fore.GREEN}✓ Escala copiada para histórico{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.YELLOW}[AVISO] Aviso ao copiar escala: {e}{Style.RESET_ALL}")
    
    print(f"\n{Fore.GREEN}{'='*60}{Style.RESET_ALL}")
    print(f"{Fore.GREEN}✓ RELATÓRIO CRIADO COM SUCESSO!{Style.RESET_ALL}")
    print(f"{Fore.CYAN}📄 Raiz: {arquivo_raiz}{Style.RESET_ALL}")
    print(f"{Fore.CYAN} Histórico: {arquivo_historico}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}📊 Escala: {arquivo_escala_destino}{Style.RESET_ALL}")
    print(f"{Fore.GREEN}{'='*60}{Style.RESET_ALL}\n")

# Main execution
if __name__ == '__main__':
    limpar_tela()
    print(f"{Fore.CYAN}{'='*60}{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{Style.BRIGHT}    GERADOR DE RELATÓRIO - OPERAÇÃO P2{Style.RESET_ALL}")
    print(f"{Fore.CYAN}{'='*60}{Style.RESET_ALL}\n")
    
    plano_do_dia = input(f"{Fore.YELLOW}📋 Qual o plano do dia? {Fore.WHITE}» {Style.RESET_ALL}").upper()
    responsavel = input(f"{Fore.YELLOW}👤 Quem é o responsável? {Fore.WHITE}» {Style.RESET_ALL}").upper()
    aguardando_mdf = input(f"{Fore.YELLOW}📦 Quantas estão aguardando MDF? {Fore.WHITE}» {Style.RESET_ALL}").upper()
    aguardando_faturamento = input(f"{Fore.YELLOW}💳 Quantas estão aguardando FATURAMENTO? {Fore.WHITE}» {Style.RESET_ALL}").upper()
    
    print(f"\n{Fore.CYAN}{'─'*60}{Style.RESET_ALL}\n")
    create_report(plano_do_dia, responsavel, aguardando_mdf, aguardando_faturamento)